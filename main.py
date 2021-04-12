# -*- coding: utf-8 -*-
"""Создание сырых netCDF файлов хромато-масс спектров.

Модуль для моделирования хромато-масс спектров на основе таких параметров пика как
время удерживания, отношения сигнал шум, параметров распределения задающего этот пик и
масс спектра для соединения пика.

Модуль позволяет провести симуляцию эксперимента, в котором параметры пиков от одной
хроматограммы к другой будут варьироваться.

Examples
--------

    $ python3 main.py -i input_xlsx.xlsx -o name_experiment

    $ python3 main.py -i input_xlsx.xlsx -o filename.cdf

    $ python3 main.py

"""

import math
import numpy as np
from scipy.stats import exponnorm
from scipy.optimize import minimize_scalar
from netCDF4 import Dataset
from copy import deepcopy
from decimal import Decimal
from typing import List, Tuple
import openpyxl as ox
from multiprocessing import cpu_count, Process, Manager
import os
import sys
import subprocess
import argparse
import time
import shutil

# CONSTS
CONST_BY_SN_SIGMA_NOISE = 4
INTENSITY_BASELINE = 15.0
CONST_BY_SQRT_NOISE = 3.8
CONST_ADDITIVE_INTENSITY_NOISE = 30.0  # обеспечивает макс-ое кол-во положительных значений с данным алгоритмом шума
STD_NOISE = CONST_BY_SN_SIGMA_NOISE * CONST_BY_SQRT_NOISE * math.sqrt(INTENSITY_BASELINE)
MORE_TO_RUN_PARALLEL = 8  # кол-во файлов в Experiment больше которого включается параллельный режим
K_EMG_TO_NORMAL = 0.001
MAX_SN = 100000
TYPE_INTENSITY = 'int'
OPEN_DIR_WHEN_CREATED = True  # автоматически открывать директорию с созданным файлом/экспериментом


def min2sec(*params, dim: str):
    """Переводит значения `params` из минут в секунды.

    Examples
    --------
    >>> a, b = 1, 2.5
    >>> a, b = min2sec(a, b, dim='min')
    >>> print(a, b)
    60 150.0
    >>> d, = min2sec(1, dim='m')
    >>> print(d)
    60
    >>> j, = min2sec(1997, dim='sec')
    >>> print(j)
    1997
    """

    if dim in ('min', 'm'):
        return map(lambda x: x * 60, params)

    elif dim in ('sec', 's'):
        return params

    raise AttributeError(f"Неизвестная размерность {dim}")


def round_index(index: float) -> int:
    """Округляет значение к ближайшему целому.

    Основное применение функции - однозначно определять конечное значение индекса.
    Отличие от стандартной функции round() в том, что округление чисел с дробной
    частью = 0.5 всегда идет в большую сторону.

    Examples
    --------
    >>> round_index(2.5)
    3
    >>> round_index(0.5)
    1
    >>> round_index(1.25)
    1
    """

    if index % 1 == 0.5:
        return math.ceil(index)

    return round(index)


def check_filename(filename: str, file_format: str) -> str:
    """Проверяет корректно ли имя файла.

    Имя файла корректно, если не содержит запрещенных символов
    и `file_format` совпадает с указанным в `filename` форматом.

    Parameters
    ----------
    filename : str
        Имя файла указанное как 'имя_файла.формат', либо просто 'имя_файла'.
    file_format : str
        Формат файла (без точки перед ним).

    Returns
    -------
    str
        Имя файла с форматом (имя_файла.формат).

    Raises
    ------
    AttributeError
        В имени файла есть запрещенные символы `\\|/*<>?:"``

    Examples
    --------
    >>> check_filename('filename', 'cdf')
    'filename.cdf'
    >>> check_filename('filename.cdf', 'cdf')
    'filename.cdf'
    >>> check_filename('filename_with_:\|<*_*>|/.cdf', 'cdf')
    Traceback (most recent call last):
        ...
    AttributeError: Имя файла не должно содержать эти символы \|/*<>?:"

    """

    # проверка на запрещенные символы в имени файла
    impossible_symbols = "\\|/*<>?:\""
    for symbol in impossible_symbols:
        if symbol in filename:
            raise AttributeError(f"Имя файла не должно содержать эти символы {impossible_symbols}")

    filename_format = filename.split('.')
    # указано только имя файла
    if len(filename_format) == 1:
        filename += f'.{file_format}'
        return filename

    # корректное имя и формат файла
    elif filename_format[-1] == file_format:
        return filename

    # имя файла корректно, а формат нет
    raise ValueError(f"Указанный формат {filename_format[-1]} файла {filename} "
                     f"не совпал с требуемым форматом {file_format}")


class Peak:
    """Хромато-масс пик.

    Класс для инициализации параметров пика и вычисления значений функции плотности
    вероятности (ФПВ) заданного распределения.

    Methods
    -------
    normal(t)
        Значение ФПВ нормального распределения в `t` - скаляр либо вектор.

    exponnormal(t)
        Значение ФПВ экспоненциально модифицированного нормального распределения (ЭМГ)
        в `t` - скаляр либо вектор.

    """

    def __init__(self, rt, sn, sigma=None, tau=None, w6=None, w05=None, mz: int = None, name: str = None,
                 mz_list=np.empty(0, dtype=np.int16), intensity_list=np.empty(0, dtype=np.float16),
                 dim='min', distribution='normal', **kwargs):
        """Инициализация параметров пика.

        Parameters
        ----------
        rt : float, int
            Время удерживания пика, в сек или мин.
        sn : float, int
            Отношение сигнал/шум для базового пика, если не передан параметр `mz`.
        sigma : float, int
            Стандартное отклонение для одного из распределений пика.
        w6 : float, int
            Ширина 6 sigma, только для нормального распределения.
        w05 : float, int
            Ширина на полувысоте, только для нормального распределения.
        tau : float, int
            Время релаксации экспоненты для ЭМГ распределения.
        mz : int, default = None
            Массовое число m/z для sn.
        name : str, default = None
            Название соединения, опционально.
        mz_list : np.ndarray of np.int16
            Массив массовых чисел m/z масс-спектра для пика.
        intensity_list : np.ndarray of np.float16
            Массив интенсивностей масс-спектра для пика.
        dim : str, default = 'min'
            Размерность времени удерживания `rt` и параметра ширины (sigma или w6 или w05), 'sec' или 'min'.
        distribution : str, default = 'normal'
            Тип распределения описывающего пик, 'normal' - нормальное (Гауссово) распределение,
            другой возможный вариант распределения 'exponnormal' - экспоненциально модифицированное
            нормальное (ЭМГ).
        """

        self.name = name
        self.distribution = distribution

        #  w -> sigma (переход от ширины к сигма параметру)
        self.sigma = sigma if sigma is not None else self.__w_to_sigma(w6=w6, w05=w05)
        assert self.sigma > 0, f'недопустимое значение sigma = {self.sigma} для {name}'

        #  min -> sec (перевод в секунды)
        self.rt, self.sigma = min2sec(rt, self.sigma, dim=dim)

        if distribution in ('normal', 'G'):
            self.max = 1
        elif distribution in ('exponnormal', 'EMG'):
            assert tau > 0, f'недопустимое значение tau = {tau} для {name}'
            self.tau, = min2sec(tau, dim=dim)
            self.k = self.tau / self.sigma
            if self.k <= K_EMG_TO_NORMAL:
                self.distribution = 'normal'
            else:
                self.mu = self.rt - self._rt_mu0
                self.max = self.exponnormal(self.rt)
        else:
            raise AttributeError(f"Неизвестное распределение {distribution}")

        self.mz_list = mz_list
        self.intensity_list = intensity_list
        self.sn = sn

        if len(intensity_list):
            self._normalization_intensity_list(type_norm='max')
            if mz is not None:
                # индекс mz для которого задано SN
                index_mz_for_sn, = np.where(mz_list == mz)[0]
                # пересчет SN для mz -> к SN для базового пика
                self.sn = sn / self.intensity_list[index_mz_for_sn]

    def normal(self, t):
        """Плотность вероятности нормального распределения в `t`.

        Функция плотности вероятности (без предэкспоненциального множителя):

        .. math::

            f(t) = \exp\left(-\frac{\left(t-\mu\right)^{2}}{2\sigma^{2}}\right)

        где :math: `\mu` положение максимума, :math: `t` время, аргумент функции,
        :math: `\sigma` - среднеквадратическое отклонение

        Parameters
        ----------
        t : float, int, np.ndarray
            Точка или вектор значений времени [сек].

        Returns
        -------
        float, np.ndarray
            Значени(е/я) функции плотности вероятности нормального распределения в `t`.
        """

        return np.exp(-(t - self.rt)**2 / (2 * self.sigma**2))

    def exponnormal(self, t):
        """Плотность вероятности экспоненциально модифицированного нормального распределения в `t`.

        Функция плотности вероятности:

        .. math::

            f(t) = \frac{1}{2K} \exp\left(\frac{1}{2K^{2}}-\frac{\frac{t-\mu}{\sigma}}{K}\right)
                   \operatorname{erfc}\left(-\frac{\frac{t-\mu}{\sigma}-\frac{1}{K}}{\sqrt{2}}\right)

        где :math:`t` время, аргумент функции; :math:`K = \frac{\tau}{\sigma} > 0`, где :math: `\tau` - время
        релаксации экспоненты, :math: '\sigma' - среднеквадратическое отклонение; :math: '\mu' - среднее
        нормального распределения.

        Parameters
        ----------
        t : float, int, np.ndarray
            Точка или вектор значений времени [сек].

        Returns
        -------
        float, np.ndarray
            Значени(е/я) функции плотности вероятности нормального распределения в `t`.
        """

        return exponnorm.pdf(t, self.k, loc=self.mu, scale=self.sigma)

    @property
    def _rt_mu0(self) -> float:
        """Время максимума для ЭМГ пика с параметром :math: `\mu = 0`.

        Применяется для нахождения положения максимума ФПВ ЭМГ с такими же параметрами
        пика `self`, за одним исключением - параметра :math: `\mu` равным нулю.
        Это значение затем используется для нахождения параметра :math: `\mu` пика `self` с известным
        положением максимума ФПВ ЭМГ.

        Notes
        -----
        Поиск минимума от отрицательных значений ФПВ ЭМГ с параметром :math: `\mu = 0`
        равносилен поиску максимума для положительных значений.

        Returns
        -------
        float
            Найденное положение максимума.
        """

        return minimize_scalar(self.__negative_exponnorm_mu0, method='brent').x

    def _normalization_intensity_list(self, type_norm='max') -> None:
        """Нормировка интенсивностей масс спектра.

        Parameters
        ----------
        type_norm : str, default 'max'
            Тип нормировки, 'max' - нормировка на максимальное значение,
            'sum' - нормировка на сумму интенсивностей.
        """

        if type_norm == 'max':
            self.intensity_list /= self.intensity_list.max()

        elif type_norm == 'sum':
            self.intensity_list /= self.intensity_list.sum()

        else:
            raise AttributeError(f"Неизвестный тип нормировки {type_norm}")

    def __negative_exponnorm_mu0(self, t: float) -> float:
        """Отрицательное значение ФПВ ЭМГ в точке `t` с параметром :math: `\mu = 0`.

        Основное применение - передача как аргумента минимизируемой функции в minimize_scalar.
        """

        return -exponnorm.pdf(t, self.k, loc=0, scale=self.sigma)

    def __w_to_sigma(self, w6=None, w05=None) -> float:

        if w6 is not None and w05 is None:
            return w6 / 6

        elif w6 is None and w05 is not None:
            return w05 / (2 * math.sqrt(2 * math.log(2)))

        raise AttributeError(f"Неизвестный тип параметра ширины")

    def __lt__(self, other) -> bool:
        return self.rt < other.rt


class Chromatogram:
    """Хроматограмма - структура данных куда помещаются хромато-масс пики.

    Основное применение класса - расположение пиков, а также создание netCDF файла получившегося
    хромато-масс спектра. Также возможно изменять параметры пиков в рамках этой хроматограммы.

    Notes
    -----
    Хроматограмма представляет собой двумерный массив, строки которого соответствуют отдельным сканам,
    а столбцы массовым числам. Значения в ячейках массива определяют интенсивность сигнала.
    Например:
                                            mz0  mz1  mz2  mz3
                                      t0   [[3,   7,   0,   2],
                                      t1    [4,   9,   0,   3],
                                      t2    [2,   5,   0,   1],
                                      t3    [1,   2,   0,   0]]
    Methods
    -------
    set_peaks(peaks)
        Установить пики на хроматограмму.

    set_noise()
        Установить шум на хроматограмму.

    replace_negative(value)
        Заменить отрицательные значения хроматограммы на `value`.

    create_cdf(filename)
        Создать netCDF файл хроматограммы.

    change_rt(times)
        Изменить времена удерживания пиков на `times`.

    change_sn(sns)
        Изменить отношения сигнал шум пиков на `sns`.

    change_shape(sigma, tau)
        Изменить форму пиков с ЭМГ распределением.

    clear()
        Очистить хроматограмму.

    update()
        Обновить хроматограмму (перерисовать) с учетом возможных изменений параметров пиков.

    update_get_peaks_areas_heights()
        Аналог update(), но с возвращением списка площадей и высот пиков.

    update_cdf(filename)
        Обновить существующий netCDF файл.

    """

    def __init__(self, t_start, t_end, mz_min: int, mz_max: int,
                 scan_rate=20, sigma_noise=1, dim_t='min', **kwargs):
        """Инициализация хроматограммы.

        Parameters
        ----------
        t_start : float or int
            Начальное время хроматограммы.
        t_end : float or int
            Время конца хроматограммы.
        mz_min : int
            Минимум оси mz.
        mz_max : int
            Максимум оси mz.
        scan_rate : int or float, default = 20
            Скорость сканирования = количество спектров в секунду.
        sigma_noise : int or float, default = 1
            Стандартное отклонение шума.
        dim_t : str, default = 'min'
            Размерность `t_start` и `t_end`.

        """

        assert scan_rate >= 1, f'Недопустимое значение scan_rate = {scan_rate}'
        assert tuple(map(int, (mz_min, mz_max))) == (
            mz_min, mz_max), f'Минимум максимум оси mz нецелые {mz_min, mz_max}'
        assert 0 <= mz_min <= mz_max, f'Недопустимые значения min max mz: {mz_min} {mz_max}'
        assert t_start <= t_end, f'Недопустимые значения start end t: {t_start} {t_end}'
        assert sigma_noise >= 0, f'Недопустимое значение sigma_noise = {sigma_noise}'
        self.sigma_noise = sigma_noise

        # float, int -> Decimal (для корректной работы с float)
        t_start, t_end, scan_rate = map(Decimal, map(str, (t_start, t_end, scan_rate)))

        # t [min] -> t [sec]
        self.t_start, self.t_end = min2sec(t_start, t_end, dim=dim_t)

        self.scan_rate = scan_rate

        self.mz_min, self.mz_max = mz_min, mz_max

        # проверка осей
        self.scans_count, self.point_count = self._get_counts_if_axes_correct

        self.data = np.zeros((self.scans_count, self.point_count), dtype=np.float64)
        self.peaks: List[Peak] = []  # для хранения пиков при их установке

        # Decimal -> float (чтобы м.б. работать с этими параметрами далее)
        self.t_start, self.t_end, self.scan_rate = map(float, (self.t_start, self.t_end, self.scan_rate))
        self.step = 1 / self.scan_rate

    def set_peaks(self, peaks: List[Peak]) -> None:
        """Размещает `peaks` на хроматограмме.

        Notes
        -----
        Для того чтобы избежать изменений исходных `peaks` создается их копия, которая доступна
        по аттрибуту хроматограммы peaks.

        Подразумевается, что этот метод используется тогда, когда хроматограмма пуста, то есть
        при самом первом размещении пиков. Повторное использование этого метода приведет к наложению
        новых пиков на старые, при этом старые пики больше не будут принадлежать этой хроматограмме,
        то есть аттрибут peaks будет соответствовать новым пикам.

        Для того, чтобы изменить параметры пиков на хроматограмме и перерисовать её, следует обращаться
        к соответствующим методам (change_rt, change_sn, change_shape и update) либо напрямую через аттрибут peaks.

        Parameters
        ----------
        peaks : List[Peak]
            Список из экземпляров класса Peak.
        """

        self.peaks = deepcopy(peaks)
        for peak in self.peaks:
            self._set_peak(peak)

    def set_noise(self, npy_file=False, experiment_name='') -> None:
        """Устанавливает гетероскедастический шум (зависит от интенсивности).

        Parameters
        ----------
        npy_file : str or bool, default = False
            Имя npy файла для сохранения матрицы шума.
        experiment_name : str, default = ''
            Имя эксперимента (варьирование параметров).

        """

        self.data += INTENSITY_BASELINE
        matrix_noise = np.random.normal(0, self.sigma_noise, size=(self.scans_count, self.point_count))
        self.data += CONST_BY_SQRT_NOISE * np.sqrt(self.data) * matrix_noise + CONST_ADDITIVE_INTENSITY_NOISE

        if npy_file:
            np.save(f'{experiment_name}/matrix_noise/{npy_file}', matrix_noise)

    def replace_negative(self, value: float) -> None:
        """Заменяет отрицательные значения хроматограммы на `value`."""
        self.data[self.data < 0] = value

    def create_cdf(self, filename: str, path='') -> None:
        """Создает netCDF файл в классическом формате netCDF3.

        Parameters
        ----------
        filename : str
            Имя cdf файла.
        path : str
            Путь по которому будет выполнено сохранение.

        Notes
        -----
        Операция превращения 2D матрицы данных (self.data) в 1D массив (intensity_values)
        представляется сл. образом:

            mz0 mz1 mz2                mz0 mz1 mz2 mz0 mz1 mz2                      mz0 mz1 mz2 mz0 mz1 mz2 mz0 mz1 mz2
        t0 [[1,  2,  3],      t0 & t1 [[1,  2,  3,  4,  5,  6],       t0 & t1 & t2 [[1,  2,  3,  4,  5,  6,  7,  8,  9]]
        t1  [4,  5,  6],  -->      t2  [7,  8,  9]]              -->
        t2  [7,  8,  9]]

        Расчет полной интенсивности (total_intensity)

            mz0 mz1 mz2                 mz0 mz1 mz2         mz0 & mz1 & mz2
        t0 [[1,  2,  3],       sum(t0) [[1 + 2 + 3],          t0 [[6],
        t1  [4,  5,  6],  -->  sum(t1)  [4 + 5 + 6],   -->    t1  [15],
        t2  [7,  8,  9]]       sum(t2)  [7 + 8 + 9]]          t2  [24]]
        """

        # Инициализируем датасет, создав/открыв `filename` на запись
        with Dataset(f"{path}/{filename}", 'w', format='NETCDF3_CLASSIC') as ds:

            # Глобальные атрибуты
            ds.dataset_completeness = "C1+C2"
            ds.ms_template_revision = "1.0.1"
            ds.experiment_date_time_stamp = ""
            ds.netcdf_file_date_time_stamp = ""
            ds.experiment_type = "Centroided Mass Spectrum"
            ds.netcdf_revision = "ncgen 3.6.1"
            ds.languages = "English"
            ds.test_separation_type = ""
            ds.test_ms_inlet = ""
            ds.test_ionization_mode = ""
            ds.test_ionization_polarity = ""
            ds.test_detector_type = ""
            ds.test_scan_function = ""
            ds.test_scan_direction = ""
            ds.test_scan_law = ""
            ds.raw_data_mass_format = "Double"
            ds.raw_data_intensity_format = "Long"
            ds.global_mass_min = float(self.mz_min)
            ds.global_mass_max = float(self.mz_max)

            # Создание измерений
            ds.createDimension('_2_byte_string', size=2)
            ds.createDimension('_4_byte_string', size=4)
            ds.createDimension('_8_byte_string', size=8)
            ds.createDimension('_16_byte_string', size=16)
            ds.createDimension('_32_byte_string', size=32)
            ds.createDimension('_64_byte_string', size=64)
            ds.createDimension('_80_byte_string', size=80)
            ds.createDimension('_128_byte_string', size=128)
            ds.createDimension('_255_byte_string', size=255)
            ds.createDimension('scan_number', size=self.scans_count)
            ds.createDimension('instrument_number', size=1)
            ds.createDimension('error_number', size=1)
            ds.createDimension('point_number', None)

            # Создание переменных
            error_log = ds.createVariable('error_log', 'S1', ('error_number', '_64_byte_string'))  # ""
            scan_index = ds.createVariable('scan_index', np.intc, 'scan_number')
            point_count = ds.createVariable('point_count', np.intc, 'scan_number')
            flag_count = ds.createVariable('flag_count', np.intc, 'scan_number')  # 0
            a_d_sampling_rate = ds.createVariable('a_d_sampling_rate', np.float64, 'scan_number')  # -9999
            scan_acquisition_time = ds.createVariable('scan_acquisition_time', np.float64, 'scan_number')
            scan_duration = ds.createVariable('scan_duration', np.float64, 'scan_number')  # 0
            mass_range_min = ds.createVariable('mass_range_min', np.float64, 'scan_number')
            mass_range_max = ds.createVariable('mass_range_max', np.float64, 'scan_number')
            resolution = ds.createVariable('resolution', np.float64, 'scan_number')  # -9999
            total_intensity = ds.createVariable('total_intensity', np.float64, 'scan_number')
            mass_values = ds.createVariable('mass_values', np.float64, 'point_number')

            # для XCMS
            instrument_name = ds.createVariable('instrument_name', 'S1', ('instrument_number', '_32_byte_string'))
            instrument_mfr = ds.createVariable('instrument_mfr', 'S1', ('instrument_number', '_32_byte_string'))

            # float64 -> int32 или float64 -> float64
            if TYPE_INTENSITY in ('int', 'i'):
                self.data = np.rint(self.data)
                intensity_values = ds.createVariable('intensity_values', np.intc, 'point_number')
            else:
                intensity_values = ds.createVariable('intensity_values', np.float64, 'point_number')

            # Создание атрибутов переменных
            total_intensity.units = 'Total Counts'
            mass_values.scale_factor = 1.0
            mass_values.units = 'M/Z'
            intensity_values.scale_factor = 1.0
            intensity_values.units = 'Arbitrary Intensity Units'

            zeros = np.zeros(self.scans_count, dtype=np.intc)  # часто используемые массивы
            const_array = np.repeat(-9999, self.scans_count)   #

            # Заполнение переменных данными
            error_log[:] = ''
            scan_index[:] = np.arange(0, self.point_count * self.scans_count, self.point_count)
            point_count[:] = np.repeat(self.point_count, self.scans_count)
            flag_count[:] = zeros
            a_d_sampling_rate[:] = const_array
            scan_acquisition_time[:] = np.linspace(self.t_start, self.t_end, num=self.scans_count, endpoint=False)
            scan_duration[:] = zeros
            mass_range_min[:] = np.repeat(self.mz_min, self.scans_count)
            mass_range_max[:] = np.repeat(self.mz_max, self.scans_count)
            resolution[:] = const_array
            total_intensity[:] = np.sum(self.data, axis=1)
            mass_values[:] = np.tile(np.arange(self.mz_min, self.mz_max + 1, 1, dtype=np.float64), self.scans_count)
            instrument_name[:] = ''
            instrument_mfr[:] = ''
            # 2D -> 1D
            intensity_values[:] = np.ravel(self.data)

    def change_rt(self, times) -> None:
        """Изменить времена удерживания пиков.

        Parameters
        ----------
        times : np.ndarray, list, tuple
            Вектор новых значений времен удерживания
            (индексация соответствует порядку перечисления пиков в `self.peaks`).
        """

        for time, peak in zip(times, self.peaks):
            peak.mu += time - peak.rt
            peak.rt = time

    def change_sn(self, sns) -> None:
        """Изменить отношения сигнал шум пиков.

        Parameters
        ----------
        sns : np.ndarray, list, tuple
            Вектор новых значений sn.
        """

        for sn, peak in zip(sns, self.peaks):
            peak.sn = sn

    def change_shape(self, sigma, tau) -> None:
        """Изменить форму ЭМГ пиков.

        Parameters
        ----------
        sigma : np.ndarray, list, tuple
            Вектор новых значений sigma.
        tau : np.ndarray, list, tuple
            Вектор новых значений tau.
        """

        for sigma, tau, peak in zip(sigma, tau, self.peaks):
            peak.sigma = sigma
            peak.tau = tau
            peak.k = tau / sigma
            peak.mu = peak.rt - peak._rt_mu0
            peak.max = peak.exponnormal(peak.rt)

    def clear(self):
        """Очистить хроматограмму."""
        self.data = np.zeros((self.scans_count, self.point_count), dtype=np.float64)

    def update(self):
        """Обновить хроматограмму (перерисовать) с учетом возможных изменений параметров пиков."""
        self.clear()
        for peak in self.peaks:
            self._set_peak(peak)

    def update_get_peaks_areas_heights(self) -> [list, list]:
        """Обновить хроматограмму и получить площади и высоты пиков.

        Отличие от метода update() в том, что этот метод помимо перерисовки хроматограммы
        вычисляет высоту и площадь пиков и затем возвращает полученный список.

        Return
        ------
        [peaks_areas, peaks_heights] : List[list, list]
            Список из списка площадей и списка высот пиков.
        """

        self.clear()
        peaks_areas = []
        peaks_heights = []
        for peak in self.peaks:
            if peak.sn <= 0:
                peaks_areas.append(0)
                peaks_heights.append(0)
                continue
            # времена (начала и конца пика) -> индексы (по оси t на хроматограмме)
            pi_start, pi_end = map(self._time2index, self._cut_peak_if_time_out(peak))
            pt_start, pt_end = map(self._index2time, (pi_start, pi_end))
            # вектор значений времени для пика (точки по оси t где расположен пик)
            time_interval = np.linspace(pt_start, pt_end, pi_end - pi_start, endpoint=False)
            mz_list, intensity_list = self._check_mz_out(peak)
            # массовые числа -> индексы
            mz_index_list = mz_list - self.mz_min

            intensity = peak.sn * STD_NOISE  # скаляр
            peaks_heights.append(intensity)

            if peak.distribution in ('normal', 'G'):
                intensity *= peak.normal(time_interval)  # вектор интенсивностей
            elif peak.distribution in ('exponnormal', 'EMG'):
                intensity *= peak.exponnormal(time_interval) / peak.max  # вектор интенсивностей

            for mz, part_intensity in zip(mz_index_list, intensity_list):
                self.data[pi_start:pi_end, mz] += intensity * part_intensity

            # расчет площади базового пика методом трапеций
            peaks_areas.append(np.trapz(y=intensity, dx=self.step))

        return [peaks_areas, peaks_heights]

    def update_cdf(self, filename: str, path=''):
        """Обновляет существующий netCDF файл."""

        with Dataset(f"{path}/{filename}", 'r+') as ds:
            if TYPE_INTENSITY in ('int', 'i'):
                self.data = np.rint(self.data)

            ds.variables['total_intensity'][:] = np.sum(self.data, axis=1)
            ds.variables['intensity_values'][:] = np.ravel(self.data)

    @property
    def _get_counts_if_axes_correct(self) -> Tuple[int, int]:
        """Вычисляет размерность хроматограммы.

        Return
        ------
        scans_count, point_counts : Tuple[int, int]
            Возвращает кортеж из целых чисел, если оси заданы корректно.
        """

        scans_count = (self.t_end - self.t_start) * self.scan_rate
        point_counts = (self.mz_max - self.mz_min) + 1

        ints = int(scans_count), int(point_counts)
        if (scans_count, point_counts) == ints:
            return ints

        raise ValueError(f"Оси/ось имеют недопустимый интервал [{self.t_start}, {self.t_end}], "
                         f"[{self.mz_min}, {self.mz_max}]")

    def _time2index(self, time: float) -> int:
        return round_index((time - self.t_start) * self.scan_rate)

    def _index2time(self, index: int) -> float:
        return index / self.scan_rate + self.t_start

    def _check_t_out(self, t_start_end: Tuple[float, float]) -> Tuple[float, float]:
        """Проверяет начало и конец пика на выход за границы временной оси.

        Parameters
        ----------
        t_start_end : Tuple[float, float]
            Кортеж из проверяемых времен начала и конца пика.

        Returns
        -------
        start, end : Tuple[float, float]
            Кортеж исправленных времен начала и конца пика.

        Raises
        ------
        ValueError
            Пик вышел за границы временной оси

        """
        start, end = t_start_end[0], t_start_end[1]

        if start < self.t_start:
            start = self.t_start
        elif start >= self.t_end:
            raise ValueError("Пик вышел за границы временной оси")

        if end <= self.t_start:
            raise ValueError("Пик вышел за границы временной оси")
        elif end > self.t_end:
            end = self.t_end - 1 / self.scan_rate

        return start, end

    def _check_mz_out(self, peak: Peak) -> Tuple[np.ndarray, np.ndarray]:
        """Проверяет значения m/z для пика на попадание в интервал оси m/z.

        Возвращает кортеж из массива массовых чисел, которые входят в интервал оси m/z
        и соответствующего этим массовым числам массива интенсивностей.
        """
        mz = peak.mz_list
        condition = (self.mz_min <= mz) & (mz <= self.mz_max)

        return mz[condition], peak.intensity_list[np.where(condition)]

    def _peak_t_start_end(self, peak: Peak) -> Tuple[float, float]:
        """Возвращает кортеж времени начала и конца пика."""

        if peak.distribution in ('normal', 'G'):

            min_intensity = (1 / 100)
            other = math.sqrt(-2 * math.log(min_intensity / peak.sn)) * peak.sigma
            return peak.rt - other, peak.rt + other

        elif peak.distribution in ('exponnormal', 'EMG'):
            min_q = 0.00000001  # lower tail probability
            max_q = 0.99999999
            return tuple(map(lambda x: exponnorm.ppf(x, K=peak.k, loc=peak.mu, scale=peak.sigma), (min_q, max_q)))

        else:
            raise AttributeError(f"Неизвестное распределение {peak.distribution}")

    def _cut_peak_if_time_out(self, peak: Peak) -> Tuple[float, float]:
        """Обрезает пик если он выходит за границы временной оси.

        Parameters
        ----------
        peak : Peak
            Экземпляр класса Peak.

        Returns
        -------
        start_end : Tuple[float, float]
            Кортеж из исправленных времен начала и конца пика.
        """

        start_end = self._peak_t_start_end(peak)
        start_end = self._check_t_out(start_end)
        return start_end

    def _set_peak(self, peak: Peak) -> None:
        """Размещает `peak` на хроматограмму.

        Установка начинается с поиска начала и конца пика. Если часть пика выходит за границы оси времени,
        то он обрезается, если же полностью располагается за границей, то возникает исключение.
        Устанавливаются только те пики, значения mz которых лежат в интервале заданной оси mz.

        Parameters
        ----------
        peak : Peak
            Экземпляр класса Peak.

        """

        # не устанавливаем пики с sn <= 0
        if peak.sn <= 0:
            return

        # времена (начала и конца пика) -> индексы (по оси t на хроматограмме)
        pi_start, pi_end = map(self._time2index, self._cut_peak_if_time_out(peak))
        pt_start, pt_end = map(self._index2time, (pi_start, pi_end))
        # вектор значений времени для пика (точки по оси t где расположен пик)
        time_interval = np.linspace(pt_start, pt_end, pi_end - pi_start, endpoint=False)

        mz_list, intensity_list = self._check_mz_out(peak)
        # массовые числа -> индексы
        mz_index_list = mz_list - self.mz_min

        intensity = peak.sn * STD_NOISE  # скаляр

        if peak.distribution in ('normal', 'G'):
            intensity *= peak.normal(time_interval)  # вектор интенсивностей

        elif peak.distribution in ('exponnormal', 'EMG'):
            intensity *= peak.exponnormal(time_interval) / peak.max  # вектор интенсивностей

        for mz, part_intensity in zip(mz_index_list, intensity_list):
            self.data[pi_start:pi_end, mz] += intensity * part_intensity

    def _fix_rt(self):
        """Исправляет времена удерживания пиков.

        Гарантируя что максимумы пиков будут зарегистрированы."""
        for peak in self.peaks:
            prev_rt = peak.rt
            peak.rt = self._index2time(self._time2index(peak.rt))
            peak.mu += peak.rt - prev_rt


class Experiment:
    """Метаболомный эксперимент.

    Применяется для создания нескольких сырых файлов netCDF, которые соответствуют одному
    метаболомному исследованию, в котором варьируются параметры пиков.

    Methods
    -------
    create_files(params_files, start_idx, list_areas_heights, noise)
        Создать netCDF файлы.

    parallel_create_files(list_areas_heights, noise)
        Создать netCDF файлы в параллельном режиме.

    write_xlsx_params()
        Записать новые параметры пиков, а также их площади и высоты в таблицу xlsx.

    """

    def __init__(self, peaks: List[Peak], chrom: Chromatogram,
                 sigma_params: dict, n_files: int, missing: list, name: str):
        """Инициализация эксперимента.

        Parameters
        ----------
        peaks : List[Peak]
            Список пиков - экземпляров класса Peak.
        chrom : Chromatogram
            Экземпляр класса Chromatogram.
        sigma_params : dict
            Словарь, ключи которого есть имена изменяемых атрибутов пиков, а значения
            являются массивами из величин сигма для нормального(rt, sigma, tau) и/или
            логнормального(sn) распределения.
        n_files : int
            Количество cdf файлов эксперимента.
        name : str
            Имя эксперимента.
        """
        assert isinstance(n_files, int) and n_files > 0, f'Число файлов должно быть положительным целым числом'
        self.peaks = deepcopy(peaks)
        self.chromatogram = deepcopy(chrom)
        self.sigma_params = sigma_params
        self.missing = missing
        self.n_files = n_files
        self.name = name
        self.params_files = self._get_params_files()  # вычисляем новые параметры для всех файлов
        self.pointer = dict(zip(self.sigma_params.keys(), range(4)))  # 4 возможных параметра варьирования
        self.__create_catalog()
        self.chromatogram.set_peaks(peaks)
        self.chromatogram.create_cdf('reference.cdf', path=f'{self.name}')  # исходный образец без варьирования и шума
        self.__create_cdf_template()  # создаем заготовки cdf файлов с общими переменными из исходного образца
        self.areas_peaks = []  # массив для хранения площадей пиков
        self.heights_peaks = []  # массив для хранения высот пиков

    def create_files(self, params_files, start_idx: int, list_areas_heights: list, noise=True):
        """Последовательно устанавливает новые параметры пиков из `params_files`
        на хроматограмме, а затем создает netCDF файл.

        Parameters
        ----------
        params_files : np.ndarray
            Массив с новыми значениями параметров пиков
        start_idx : int
            Начальный индекс массива `params_files`. Необходим для возможности в параллельном режиме
            корректно давать имена файлам в соответствии с порядком перечисления их в `params_files`.
        list_areas_heights : list
            Список для хранения площадей и высот пиков и номера файла [[<номер файла>, [[<площади>], [<высоты>]]],..,].
        noise : bool, default = True
            Создать файлы с шумом.

        Notes
        -----
        Для эксперимента с 13 пиками и хроматограммы размером 600 х 450 (длительность хроматограммы 30 сек,
        scan_rate = 20, mz от 50 до 500) - следует использовать этот метод, только если количество файлов <= 10.
        При большем количестве файлов (или увеличении хроматограммы и/или количества пиков) эффективнее
        использовать параллельную версию этого метода - parallel_create_files().
        """

        # вариант варьирования rt & sn & sigma & tau с шумом
        if 'rt' in self.pointer and 'sn' in self.pointer and \
                'sigma' in self.pointer and 'tau' in self.pointer and noise:
            rt_idx = self.pointer['rt']
            sn_idx = self.pointer['sn']
            sigma_idx = self.pointer['sigma']
            tau_idx = self.pointer['tau']

            for i, file in enumerate(params_files, start_idx + 1):
                # извлекаем параметры по столбцам
                times = file[:, rt_idx]
                sns = file[:, sn_idx]
                sigma = file[:, sigma_idx]
                tau = file[:, tau_idx]

                # изменяем параметры пиков на хроматограмме
                self.chromatogram.change_shape(sigma, tau)
                self.chromatogram.change_rt(times)
                self.chromatogram.change_sn(sns)

                # переустанавливаем пики и вычисляем их площади и высоты
                peaks_areas_heights = self.chromatogram.update_get_peaks_areas_heights()
                list_areas_heights.append([i, peaks_areas_heights])

                self.chromatogram.set_noise(f'mn_{i}.npy', self.name)  # матрица шума сохраняется
                self.chromatogram.replace_negative(0)
                self.chromatogram.update_cdf(f'{i}.cdf', path=f'{self.name}/output_cdf')

        # вариант варьирования rt & sn & sigma & tau без шума
        elif 'rt' in self.pointer and 'sn' in self.pointer and \
                'sigma' in self.pointer and 'tau' in self.pointer and not noise:
            rt_idx = self.pointer['rt']
            sn_idx = self.pointer['sn']
            sigma_idx = self.pointer['sigma']
            tau_idx = self.pointer['tau']

            for i, file in enumerate(params_files, start_idx + 1):
                # извлекаем параметры по столбцам
                times = file[:, rt_idx]
                sns = file[:, sn_idx]
                sigma = file[:, sigma_idx]
                tau = file[:, tau_idx]

                # изменяем параметры пиков на хроматограмме
                self.chromatogram.change_shape(sigma, tau)
                self.chromatogram.change_rt(times)
                self.chromatogram.change_sn(sns)

                # переустанавливаем пики и вычисляем их площади и высоты
                peaks_areas_heights = self.chromatogram.update_get_peaks_areas_heights()
                list_areas_heights.append([i, peaks_areas_heights])
                self.chromatogram.update_cdf(f'{i}.cdf', path=f'{self.name}/output_cdf')

    def parallel_create_files(self, list_areas_heights: list, noise=True):
        """Запуск метода create_files на нескольких процессах.

        Массив с новыми параметрами пиков делится на n частей, где n - кол-во ядер процессора.
        Запущенные n процессов выполняют метод create_files для полученной части массива.

        Parameters
        ----------
        list_areas_heights : list
            Список для хранения площадей и высот пиков и номера файла [[<номер файла>, [[<площади>], [<высоты>]]],..,].
        noise : bool, default = True
            Установка шума.
        """

        parts = np.array_split(self.params_files, cpu_count())  # массив из частей

        # ищем индексы полученных частей в исходном массиве
        acc = 0
        idxs_end = []
        for split in parts:
            acc += len(split)
            idxs_end.append(acc)
        idxs_start = [0] + idxs_end[:-1]

        # создаем и запускаем процессы
        procs = []
        for idxs, part in zip(idxs_start, parts):
            proc = Process(target=self.create_files, args=(part, idxs, list_areas_heights, noise))
            procs.append(proc)
            proc.start()

        for proc in procs:
            proc.join()

    def write_xlsx_params(self) -> None:
        """Записывает параметры пиков для каждого файла в единый xlsx файл."""

        wb = ox.Workbook()
        ws = wb.active
        ws.title = 'output'

        # конечная таблица построчно
        table = [[None, ],
                 ['Peak names', ]]

        # добавляем имена пиков
        table += [[peak.name] for peak in self.peaks]
        variation_params = list(self.sigma_params.keys()) + ['Area'] + ['Height']
        prefix_filename = ''
        n_files = self.n_files
        num_var_params = len(variation_params) - 1

        # имена файлов
        for i in range(1, n_files + 1):
            table[0].extend([f'{prefix_filename}{i}.cdf'] + [None] * num_var_params)
        # строка с именами параметров для файлов
        table[1].extend(variation_params * n_files)

        # записываем значения новых параметров пиков для файлов
        cc_table = [[] for _ in range(len(self.peaks))]
        for i, file in enumerate(self.params_files):
            cc_table = np.concatenate((cc_table, file, self.areas_peaks[i].T, self.heights_peaks[i].T), axis=1)
        table_column = cc_table.tolist()

        for p, row in enumerate(table_column):
            table[2 + p] += row

        for r in table:
            ws.append(r)

        # объединяем ячейки с именами файлов
        s = 2
        e = s + num_var_params
        for i in range(1, n_files + 1):
            ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
            s = e + 1
            e = s + num_var_params

        wb.save(f'{self.name}/info_output.xlsx')

    def _get_params_files(self) -> np.ndarray:
        """Вычисляет новые параметры пиков, используя словарь sigma_params, для будущих файлов.

        Returns
        -------
        np.ndarray
            Массив столбцы которого соответствуют новым значениям параметров, а строки пику.
        """

        num_peaks = len(self.peaks)

        # извлекаем начальные параметры
        ref_params = {}
        for key in self.sigma_params.keys():
            ref_params.update({key: np.array([eval(f'peak.{key}') for peak in self.peaks])})

        # вычисляем новые параметры на основе начальных
        new_params = {}
        for key, sigma_array in self.sigma_params.items():
            if key == 'sn':
                new_params.update({key: np.random.lognormal(np.log(ref_params[key]), sigma_array,
                                                            size=(self.n_files, num_peaks))})
            else:
                new_params.update({key: ref_params[key] + np.random.normal(0, sigma_array,
                                                                           size=(self.n_files, num_peaks))})
        # заменяем выходящие за макс. sn на допустимый максимум
        new_params['sn'][new_params['sn'] > MAX_SN] = MAX_SN

        # учитываем пропуски (Missing)
        indexes_files = np.arange(self.n_files)
        for idx_peak, part in enumerate(self.missing):
            if part > 0:
                count_files = round(part * self.n_files)  # кол-во файлов с sn=0
                np.random.shuffle(indexes_files)  # случайно перемешиваем индексы файлов
                fsh_indexes_files = indexes_files[:count_files]  # выбираем первые неск. файлов с sn=0

                for idx_file in fsh_indexes_files:
                    new_params['sn'][idx_file][idx_peak] = 0

        params_files = np.stack(tuple(sigma_array for sigma_array in new_params.values()), axis=-1)

        return params_files

    def __create_catalog(self):
        # создаем каталог с экспериментом (если такой уже существует будет перезаписан!)
        if os.path.exists(self.name):
            shutil.rmtree(self.name)  # удаляем существующий каталог
        # записываем новый
        os.makedirs(f"{self.name}/matrix_noise", exist_ok=True)
        os.makedirs(f"{self.name}/output_cdf", exist_ok=True)

    def __create_cdf_template(self):

        for i in range(1, self.n_files + 1):
            shutil.copy(f'{self.name}/reference.cdf', f'{self.name}/output_cdf/{i}.cdf')


def import_xlsx(filename, skip_variation=False):
    """Импортирует параметры пиков, хроматограммы и варьируемых параметров, если они указаны.

    Parameters
    ----------
    filename : str
        Имя xlsx файла.
    skip_variation : bool, default = False
        Пропустить блок Variation даже если он есть.

    Returns
    -------
    Tuple[List[Peak], Chromatogram, dict, int, np.ndarray]
        Если в xlsx файле есть блок Variation, то вернется кортеж, в который входит список из
        экземпляров класса Peak, экземпляр класса Chromatogram, словарь варьируемых параметров
        со списками сигм, количество файлов и массив с долями файлов в которых будут пропущенны пики.

    Tuple[List[Peak], Chromatogram]
        Если же блок Variation отсутствует, то вернется кортеж только из списка экземпляров класса
        Peak и экземпляра класса Chromatogram.
    """

    wb = ox.load_workbook(filename, data_only=True)
    sheet_ranges = wb['input']
    max_row = sheet_ranges.max_row
    rows = list(sheet_ranges.rows)
    wb.close()

    def get_row(row, key):
        return list(map(lambda x: x.value, rows[row][d_xl[key]['start_idx']:
                                                     d_xl[key]['end_idx']]))

    def get_col(col, start_row, nn):
        res = []
        for i_cell in range(start_row, start_row + nn):
            res.append(sheet_ranges.cell(i_cell, col).value)
        return res

    d_xl = {}

    # читаем первую строку
    for cell in rows[0]:
        cell_value = cell.value
        if cell_value is not None:
            d_xl.update({cell_value: {}})

    # обработка объединенных ячеек (Chromatogram, Peaks, Variation)
    mcr = sheet_ranges.merged_cells.ranges
    for cr in mcr:
        name = cr.start_cell.value
        if name in d_xl:
            start_idx = cr.start_cell.col_idx - 1
            cols = cr.size['columns']
            end_idx = start_idx + cols
            d_xl[name].update({'start_idx': start_idx, 'cols': cols, 'end_idx': end_idx})

    # Chromatogram
    names, values = map(lambda x: get_row(x, 'Chromatogram'), (1, 2))
    d_xl['Chromatogram'].update(zip(names, values))
    chrom = Chromatogram(**d_xl['Chromatogram'])

    # Peaks
    head_peaks = get_row(1, 'Peaks')
    params_peak = {}
    sep_mz_i = ';'
    sep_into_mz_i = ' '
    peak_list = []
    for i in range(2, max_row):
        params_peak.update(zip(head_peaks, get_row(i, 'Peaks')))
        mz_i = np.fromstring(params_peak['mass_spect'].replace('\n', '').
                             replace(sep_mz_i, ''), sep=sep_into_mz_i).reshape((-1, 2))
        del params_peak['mass_spect']
        mz_list = mz_i[:, 0].astype(np.int16)

        peak_list.append(Peak(mz_list=mz_list, intensity_list=mz_i[:, 1], **params_peak))

    # Variation
    if 'Variation' in d_xl and not skip_variation:
        head_variation = get_row(1, 'Variation')
        params_variation = {}
        for par in head_variation:
            params_variation.update({par: []})
        for i in range(2, max_row):
            for key, value in zip(head_variation, get_row(i, 'Variation')):
                params_variation[key].append(value)
        num_files = 0
        for n, i in enumerate(rows[0]):
            if i.value in ('Num_files', 'Num files'):
                num_files = rows[1][n].value
                break

        # Missing
        miss = np.zeros(max_row)
        for n, i in enumerate(rows[0]):
            if i.value in ('Missing', 'missing', 'miss'):
                miss = np.array(get_col(n + 1, 3, len(peak_list)))
                break

        return peak_list, chrom, params_variation, num_files, miss

    return peak_list, chrom


def main_experiment(xlsx: str, experiment_name: str, noise=True):
    """Создание файлов метаболомного эксперимента `experiment_name`.

    Каталог с данными эксперимента создается в той же директории, откуда запускается
    main.py и устроен сл. образом:

    ├── experiment_name          каталог с данными эксперимента
    │ ├── matrix_noise               каталог с сохраненными матрицами шума в формате .npy
    │ ├── output_cdf                 каталог с cdf файлами эксперимента
    │ ├── info_output.xlsx           xlsx таблица с новыми параметрами для каждого файла из output_cdf
    │ ├── input_xlsx.xlsx            копия xlsx таблицы `xlsx` с исходными параметрами эксперимента
    │ └── reference.cdf              cdf файл с параметрами из `xlsx` без вариации и без шума
    │
    ├── input_xlsx.xlsx          входящая xlsx таблица с параметрами эксперимента
    ├── main.py                  мы здесь (файл с исполняемым кодом)

    Parameters
    ----------
    xlsx : str
        Имя .xlsx файла из которого будут взяты параметры для эксперимента
        (из соответствующих блоков/столбцов таблицы).
    experiment_name : str
        Имя каталога в котором будут расположены все файлы эксперимента.
    noise : bool, default = True
        Включает установку шума.
    """

    print(f'-> Создание эксперимента {experiment_name}')
    # Инициализация эксперимента, используя параметры импортируемые из xlsx файла
    experiment = Experiment(*import_xlsx(xlsx), name=experiment_name)
    # создаем копию xlsx таблицы с исходными параметрами в каталог эксперимента
    shutil.copy(xlsx, f'{experiment_name}/input_xlsx.xlsx')

    # Создание файлов эксперимента
    if experiment.n_files >= MORE_TO_RUN_PARALLEL and cpu_count() > 1:
        print(f"-> Параллельное создание {experiment.n_files} файлов...")
        _peaks_areas_heights = Manager().list()  # proxy-лист для общего использования между запущенными процессами
        experiment.parallel_create_files(list_areas_heights=_peaks_areas_heights, noise=noise)
        _peaks_areas_heights.sort()  # сортируем по номерам файлов
        peaks_areas_heights = deepcopy(_peaks_areas_heights)  # proxy-list -> list, для возможности изменения списка
    else:
        print(f"-> Последовательное создание {experiment.n_files} файлов...")
        peaks_areas_heights = []
        experiment.create_files(experiment.params_files, 0, list_areas_heights=peaks_areas_heights, noise=noise)

    # Обработка массива с площадями и высотами пиков и запись параметров пиков всех файлов в xlsx файл
    for file in peaks_areas_heights:
        del file[0]  # удаляем номера файлов из списка
        # записываем площади и высоты пиков в соответствующий аттрибут эксперимента
        experiment.areas_peaks.append([file[0][0]])
        experiment.heights_peaks.append([file[0][1]])
    # превращаем полученные массивы в np.ndarray для возможности транспонирования
    experiment.areas_peaks = np.array(experiment.areas_peaks)
    experiment.heights_peaks = np.array(experiment.heights_peaks)
    experiment.write_xlsx_params()  # создание info_output.xlsx

    # Вывод сообщений и открытие директории с экспериментом
    dir_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '', experiment.name))
    print(fr"-> Эксперимент {experiment.name} успешно создан и его данные находятся в {dir_path}")
    if OPEN_DIR_WHEN_CREATED:
        if sys.platform == 'win32':
            subprocess.Popen(f'explorer {dir_path}')
        elif sys.platform == 'linux':
            subprocess.call(['xdg-open', dir_path])
        else:
            pass  # не открываем директорию с созданным экспериментом на других платформах


def main_one_file(xlsx: str, cdf_filename: str, noise=True):
    """Создает один файл с параметрами из таблицы `xlsx`.

    Файл будет создан в той же директории откуда запускается main.py.

    ├── out_file.cdf             созданный файл
    ├── input_xlsx.xlsx          входящая xlsx таблица с параметрами файла
    ├── main.py                  мы здесь (файл с исполняемым кодом)

    Parameters
    ----------
    xlsx : str
        Имя .xlsx файла из которого будут взяты параметры для файла (блок с вариациями
        будет игнорироваться).
    cdf_filename : str
        Имя cdf файла.
    noise : bool, default = True
        Включает установку шума.
    """

    print(f'-> Создание файла {cdf_filename}')
    peaks, chromatogram = import_xlsx(xlsx, skip_variation=True)
    chromatogram.set_peaks(peaks)
    if noise:
        chromatogram.set_noise()
        chromatogram.replace_negative(0)

    dir_path = os.path.dirname(os.path.realpath(__file__))
    filename = check_filename(cdf_filename, 'cdf')
    chromatogram.create_cdf(filename=filename, path=rf'{dir_path}')
    print(f"-> Файл {cdf_filename} успешно создан и находится в {dir_path}")
    if OPEN_DIR_WHEN_CREATED:
        if sys.platform == 'win32':
            subprocess.Popen(f'explorer {dir_path}')
        elif sys.platform == 'linux':
            subprocess.call(['xdg-open', dir_path])
        else:
            pass  # не открываем директорию с созданным файлом на других платформах


def parse_args():
    """Читает параметры переданные при запуске из терминала (командной строки).

    Получает из командной строки имя xlsx файла с параметрами и имя эксперимента
    (так будет назван каталог с данными) или имя файла (с указанием формата `.cdf`).

    """

    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--import-xlsx', dest='i', type=str, help='Имя xlsx файла')
    parser.add_argument('-o', '--output', dest='o', type=str, help='Имя файла/эксперимента')
    parser.add_argument('-dn', '--disable-noise', dest='dn', action='store_true', help='Генерация без шума')

    args = parser.parse_args()

    # эксперимент
    if args.i and not args.o.split('.')[-1] == 'cdf':
        try:
            if args.dn:
                print(f'-> Установка шума отключена')
                main_experiment(args.i, args.o, noise=False)
            else:
                main_experiment(args.i, args.o, noise=True)
        except Exception as e:
            raise e
        return True
    # один файл
    elif args.i and args.o.split('.')[-1] == 'cdf':
        try:
            if args.dn:
                print(f'-> Установка шума отключена')
                main_one_file(args.i, args.o, noise=False)
            else:
                main_one_file(args.i, args.o, noise=True)
        except Exception as e:
            raise e
        return True


if __name__ == '__main__':

    start = time.time()
    # обработка параметров из командной строки
    if parse_args():
        pass
    # иначе попробовать сгенерировать из стандартного файла в стандартный вывод
    else:
        print(f'-> Не были переданы параметры в командной строке')
        try:
            print(f'-> Попытка создать эксперимент из стандартного .xlsx файла...')
            main_experiment('input_xlsx.xlsx', 'out_experiment')
        except TypeError:
            print(f'-> Неудачно')
            try:
                print(f'-> Попытка создать файл из стандартного .xlsx файла...')
                main_one_file('input_xlsx.xlsx', 'out_file.cdf')
            except Exception as e:
                print(f'-> Неудачно')
                raise e
    print(f'-> Время выполнения: {(time.time() - start):.2f} сек')
